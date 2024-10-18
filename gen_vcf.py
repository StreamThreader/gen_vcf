import os
import openpyxl
import codecs
import re
import sys
from pathlib import Path

# version: 1.5

if len(sys.argv) == 1:
    print("Full vcard, use: ./gen_vcf.py filename.xlsx")
    print("Only one login vcard, use: ./gen_vcf.py "+
          "filename.xlsx userlogin\n")
    exit(1)

# file name argument
if len(sys.argv) >= 2:
    xlsx_file_name = str(sys.argv[1])
    xlsx_file = Path(xlsx_file_name)
    if not os.path.isfile(xlsx_file_name):
        print("File: "+xlsx_file_name+" not found\n")
        exit(1)

# by default generate full vcard
select_login = "all"
# if argument given
if len(sys.argv) >= 3:
    select_login = str(sys.argv[2])

# default language (RU, UK, EN)
def_lang = "RU"
# Define def lang index
if def_lang == "UK":
    lang_last_name = 5
    lang_first_name = 6
    lang_middle_name = 7
elif def_lang == "EN":
    lang_last_name = 9
    lang_first_name = 10
    lang_middle_name = 11
else:
    lang_last_name = 1
    lang_first_name = 2
    lang_middle_name = 3


# load xlsx master-file, and open active sheet
wb_object = openpyxl.load_workbook(xlsx_file)
sheet = wb_object.active

# get max rows
max_row = sheet.max_row

# get max columns
max_col = sheet.max_column

# init list of empty dictionaries
# for parse each row to individual dict
contact_obj = [{
    "last_name":"",
    "first_name":"",
    "middle_name":"",
    "full_name":"",
    "login":"",
    "description":"",
    "email":"",
    "phone":"",
    "active":"",
    "org":"",
    "photo":"",
    "birthday":"",
    "gender":""
} for i in range(max_row)]


# go trought all row
row_index=0
for row in sheet.iter_rows():
    column_index=1

    # get each cell from current row
    for cell in row:
        # skip empty cell
        if cell.value == None:
            column_index += 1
            continue

        # parse cell by column        
        if column_index == lang_last_name:
            contact_obj[row_index]["last_name"] = cell.value
        elif column_index == lang_first_name:
            contact_obj[row_index]["first_name"] = cell.value
        elif column_index == lang_middle_name:
            contact_obj[row_index]["middle_name"] = cell.value
        elif column_index == 13:
            contact_obj[row_index]["login"] = cell.value
        elif column_index == 15:
            contact_obj[row_index]["description"] = cell.value
        elif column_index == 16:
            contact_obj[row_index]["email"] = cell.value
        elif column_index == 17:
            contact_obj[row_index]["phone"] = cell.value
        elif column_index == 18:
            contact_obj[row_index]["active"] = cell.value
        elif column_index == 19:
            contact_obj[row_index]["org"] = cell.value
        elif column_index == 20:
            contact_obj[row_index]["photo"] = cell.value
        elif column_index == 21:
            contact_obj[row_index]["birthday"] = str(cell.value)
        elif column_index == 22:
            contact_obj[row_index]["gender"] = str(cell.value)
        column_index += 1
    row_index += 1
    # current row parsed, go next

# close xlsx
wb_object.close()

# open vcf file fow write
vcf_file = codecs.open("vcard.vcf", "w", "utf-8")


# go throught all list
for row_index in range(max_row):

    # if login given as argument
    if (select_login == "all" or
        select_login == contact_obj[row_index]["login"]):

        # skip first header row
        if row_index == 0:
            continue

        # skip row if last_name cell empty at all
        if contact_obj[row_index]["last_name"] == "":
            continue

        vcf_file.write("BEGIN:VCARD\n")
        vcf_file.write("VERSION:3.0\n")

        vcf_file.write("FN:"+contact_obj[row_index]["last_name"]+
              " "+contact_obj[row_index]["first_name"]+
              " "+contact_obj[row_index]["middle_name"]+"\n")

        vcf_file.write("N:"+contact_obj[row_index]["last_name"]+
              ";"+contact_obj[row_index]["first_name"]+
              ";"+contact_obj[row_index]["middle_name"]+";;\n")

        # split multiple emails, skip empty
        for email_addr in re.split(r"\s+",
                                   contact_obj[row_index]["email"]):
            if email_addr == "":
                continue
            vcf_file.write("EMAIL;TYPE=work:"+email_addr+"\n")

        # split multiple phone numbers, skip empty and t-escape char
        for phone_num in re.split(r"\s+",
                                  contact_obj[row_index]["phone"]):
            if phone_num == "t":
                continue
            if phone_num == "":
                continue
            vcf_file.write("TEL;TYPE=WORK:"+phone_num+"\n")

        vcf_file.write("NOTE:"+
                       contact_obj[row_index]["description"]+"\n")
        vcf_file.write("ORG:"+contact_obj[row_index]["org"]+"\n")

        # if empty, not write
        if contact_obj[row_index]["photo"] != "":
            vcf_file.write("PHOTO;ENCODING=b;TYPE=JPEG:"+
                           contact_obj[row_index]["photo"]+"\n")

        # if empty, not write
        if contact_obj[row_index]["birthday"] != "":
            vcf_file.write("BDAY:"+
                           contact_obj[row_index]["birthday"]+"\n")

        # if empty, not write
        if contact_obj[row_index]["gender"] != "":
            vcf_file.write("GENDER:"+
                           contact_obj[row_index]["gender"]+"\n")

        vcf_file.write("END:VCARD\n")

        # break if found selected login
        # already writed to vcard
        if select_login == contact_obj[row_index]["login"]:
            break
    else:
        # if reach end without result
        if row_index == (max_row - 1):
            print("not found login: "+select_login+"\n")


vcf_file.close()

print("done")



